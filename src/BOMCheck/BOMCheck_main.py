import time
from openpyxl import load_workbook
from os import rename
from src.Selenium import Selenium


class IhsMarkit:

    def __init__(self, report):
        self.wb = load_workbook(report)
        self.input_sheet = self.wb["CES"]           # hier muss man die Name von sheet eingeben
        self.max_row = self.input_sheet.max_row
        self.ihs = Selenium.ChromeIhs()
        self.part_number_list = []
        self.ordering_text_list = [28, 29, 30]          # das Programm sucht nur die Zeillennummer in diesem List.
        self.ordering_text_col = 30                 # das ist Ordering Text colummnnummer
        self.manufacturer_id_col = 29               # das ist Manufacturer ID colummnnummer
        self.part_number_col = 3                    # das ist Part number colummnnummer
        self.ext_datatabase_manu_col = 39           # das ist external database manufacturer colummnnummer

        #for row in self.ordering_text_list:
        for row in range(5, self.max_row+1):      # wenn man bestimte range von  Zeilen suchen will, dann muss den range geändert werden, range(von , bis ).
            print("current row is ####################################  " + str(row))
            # if self.input_sheet.cell(row=row, column=40).value != "Found":
            self.manufacturer_id = self.input_sheet.cell(row=row, column=self.manufacturer_id_col).value
            self.part_number1 = self.input_sheet.cell(row=row, column=self.part_number_col).value
            self.ordering_text_input = self.input_sheet.cell(row=row, column=self.ordering_text_col).value
            self.ordering_text = check_input_ordering_text1(self.ordering_text_input)
            self.ext_manufacturer_id1 = self.input_sheet.cell(row=row, column=self.ext_datatabase_manu_col).value
            self.ext_manufacturer_id = check_input_ordering_text(self.ext_manufacturer_id1)
            first_word = str(self.ext_manufacturer_id).split()
            first_word = first_word[0]
            self.ext_manufacturer_id = str(first_word)
            print(self.ordering_text)
            print(self.ext_manufacturer_id)
            self.table_list = self.ihs.get_ihs_table_list(self.ordering_text)
            self.table_elements = len(self.table_list)

            if self.table_elements == 0:
                self.input_sheet.cell(row=row, column=40).value = "Not Found"

            elif self.table_elements != 0:
                for element in range(0, self.table_elements):
                    try:
                        if "-" in self.table_list[element]:
                            self.table_list[element] = self.table_list[element].replace("-", "")
                        if str(self.table_list[element]) == str(self.ordering_text) and str(self.ext_manufacturer_id).upper() in str(self.table_list[element+1]).upper():
                            self.input_sheet.cell(row=row, column=40).value = "Found"
                            current_row = int(element/5 + 1)
                            self.Status = self.ihs.get_status_value(current_row=current_row)
                            self.result = self.ihs.eu_rohs_exemptions()
                            self.input_sheet.cell(row=row, column=42).value = self.Status
                            self.input_sheet.cell(row=row, column=41).value = self.result
                            #self.ihs.download_print()
                            # self.ihs.download_pdf_ihs(current_row)
                            if self.ihs.year > 2014:
                                self.rename_pdf()
                            break
                        else:
                            self.input_sheet.cell(row=row, column=40).value = "Hersteller nicht gefunden"
                    except:
                        print("Fehler beim {}  {} ".format(self.ordering_text, self.ext_manufacturer_id))
                        self.ihs.refresh()
                        time.sleep(2)

        self.wb.save(report)

    def rename_pdf(self):
        self.part_number_list.append(self.part_number1)
        count = self.part_number_list.count(self.part_number1)
        if count == 1:
            try:
                rename(r"C:\Temp\test_bomcheck\download.pdf".format(self.part_number1),
                       r"C:\Temp\test_bomcheck\{}_{}.pdf".format(self.part_number1, self.manufacturer_id))
            except FileExistsError:
                print("pdf ist schon vorhanden")
        else:
            try:
                rename(r"C:\Temp\test_bomcheck\download.pdf".format(self.part_number1),
                       r"C:\Temp\test_bomcheck\{}_{}({}).pdf".format(self.part_number1, self.manufacturer_id, count - 1))
            except FileExistsError:
                print("pdf ist schon vorhanden")


def check_input_ordering_text(ordering_text):
    if ">" in ordering_text:
        ordering_text = ordering_text[ordering_text.find('>'):]
        ordering_text = ordering_text.replace(">", "")
    if "/" in ordering_text:
        separator = "/"
        ordering_text = ordering_text.split(separator, 1)[0]
    return ordering_text


def check_input_ordering_text1(ordering_text):
    if ">" in ordering_text:
        ordering_text = ordering_text[ordering_text.find('>'):]
        ordering_text = ordering_text.replace(">", "")
    if "-" in ordering_text:
        ordering_text = ordering_text.replace("-", "")

    return ordering_text


def change_pdf_name(part_number, manufacturer_id, part_number1, count):
    if count == 1:
        rename(r"C:\Temp\test_bomcheck\Declaration_{}.pdf".format(part_number),
               r"C:\Temp\test_bomcheck\{}_{}.pdf".format(part_number1, manufacturer_id))
    else:
        rename(r"C:\Temp\test_bomcheck\Declaration_{}.pdf".format(part_number),
               r"C:\Temp\test_bomcheck\{}_{}({}).pdf".format(part_number1, manufacturer_id, count-1))


def test():
    rename(r"C:\Temp\test_bomcheck\download.pdf", r"C:\Temp\test_bomcheck\10816763_VIT.pdf")


def main(report):

    IhsMarkit(report)
    # test()


if __name__ == "__main__":
    file = r"C:\bomcheck\BOMCheck_test1.xlsx"   # Hier muss man die Dateipfad eingeben.
    main(file)
